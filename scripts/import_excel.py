"""
Script mejorado para importar datos del Excel a la aplicación web.
Lee estudiantes, clases y registros diarios de chill-outs.
Ejecutar con: python scripts/import_excel.py
"""

import pandas as pd
import openpyxl
import json
import sys
import os
import re
from datetime import datetime

# Agregar el directorio raíz al path
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

def import_students(filename):
    """Importa estudiantes de la hoja MAESTRA STUDENTEN"""
    try:
        wb = openpyxl.load_workbook(filename, data_only=True)
        ws = wb["MAESTRA STUDENTEN"]
        
        students = []
        header_row = None
        
        # Buscar la fila con headers (Klas, Student, Status) - está en la fila 4
        for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
            row_str = ' '.join([str(cell) if cell else '' for cell in row])
            if 'Klas' in row_str and ('Student' in row_str or 'Naam' in row_str):
                header_row = row_idx
                break
        
        if header_row is None:
            print("No se encontró la fila de headers, usando fila 4 por defecto")
            header_row = 4
        
        # Leer estudiantes desde la fila siguiente al header
        student_id_counter = 1
        for row_idx in range(header_row + 1, ws.max_row + 1):
            klas_cell = ws.cell(row=row_idx, column=1).value
            name_cell = ws.cell(row=row_idx, column=2).value
            status_cell = ws.cell(row=row_idx, column=3).value
            
            if klas_cell and name_cell:
                klas = str(klas_cell).strip()
                name = str(name_cell).strip()
                
                # Saltar filas vacías o inválidas
                if klas.lower() in ['nan', 'none', ''] or name.lower() in ['nan', 'none', '']:
                    continue
                
                # Saltar si es el header mismo
                if klas.lower() == 'klas' or name.lower() == 'student':
                    continue
                
                status = 'Actief'
                if status_cell:
                    status_str = str(status_cell).strip()
                    if status_str in ['Actief', 'Inactief', 'Verwijderd']:
                        # Si está "Verwijderd", marcarlo como Inactief
                        status = 'Inactief' if status_str == 'Verwijderd' else status_str
                
                students.append({
                    'id': f'imported_{student_id_counter}_{abs(hash(name + klas)) % 100000}',
                    'name': name,
                    'klas': klas,
                    'status': status
                })
                student_id_counter += 1
        
        wb.close()
        return students
    except Exception as e:
        print(f"Error importando estudiantes: {e}")
        import traceback
        traceback.print_exc()
        return []

def parse_chillout_value(cell_value):
    """Parsea el valor de una celda de chill-out a formato de la app"""
    if not cell_value or pd.isna(cell_value):
        return None
    
    cell_str = str(cell_value).strip()
    
    # Si está vacío
    if cell_str == '' or cell_str.lower() == 'nan':
        return None
    
    # Buscar patrones como "1 VR", "2 VL", etc.
    match = re.match(r'(\d+)\s*(VR|VL)', cell_str, re.IGNORECASE)
    if match:
        count = int(match.group(1))
        tipo = match.group(2).upper()
        if count >= 1 and count <= 3 and tipo in ['VR', 'VL']:
            return {
                'count': count,
                'type': tipo
            }
    
    return None

def find_student_in_sheet(ws, student_name, start_row, end_row):
    """Encuentra la fila de un estudiante en la hoja"""
    for row_idx in range(start_row, end_row + 1):
        name_cell = ws.cell(row=row_idx, column=1).value
        if name_cell and str(name_cell).strip().lower() == student_name.lower():
            return row_idx
    return None

def import_daily_records(filename, students_map):
    """Importa registros diarios de las hojas de días"""
    try:
        wb = openpyxl.load_workbook(filename, data_only=True)
        excel_file = pd.ExcelFile(filename)
        
        daily_sheets = [s for s in excel_file.sheet_names 
                       if any(char.isdigit() for char in s) 
                       and len(s) < 15 
                       and 'week' not in s.lower()
                       and 'totaal' not in s.lower()]
        
        records = {}
        
        for sheet_name in daily_sheets:
            try:
                # Extraer fecha del nombre de la hoja (formato: "01-09 Ma")
                parts = sheet_name.split()
                if len(parts) < 1:
                    continue
                
                date_part = parts[0]  # "01-09"
                try:
                    day, month = date_part.split('-')
                    # Determinar año basado en la fecha
                    current_year = datetime.now().year
                    # Si el mes es septiembre o posterior y estamos antes de septiembre, usar año anterior
                    if int(month) >= 9:
                        year = 2024 if datetime.now().month < 9 else 2025
                    else:
                        year = 2025
                    
                    date_str = f"{year}-{month}-{day.zfill(2)}"
                except:
                    print(f"Error parseando fecha de {sheet_name}")
                    continue
                
                ws = wb[sheet_name]
                day_name = parts[1] if len(parts) > 1 else ''
                
                # Crear registro para este día
                daily_record = {
                    'date': date_str,
                    'dayName': day_name,
                    'entries': {}
                }
                
                # Buscar secciones de clases
                for row_idx in range(1, min(ws.max_row + 1, 200)):
                    klas_cell = ws.cell(row=row_idx, column=1).value
                    if klas_cell:
                        klas_str = str(klas_cell).strip()
                        # Buscar patrones de clase como "1 Aarde", "2 Vuur", etc.
                        if re.match(r'^\d+(\/\d+)?\s+[A-Za-z]', klas_str):
                            # Encontrar estudiantes de esta clase
                            klas_students = [s for s in students_map.values() if s['klas'] == klas_str]
                            if klas_students:
                                # Buscar la fila "Naam?" que sigue al título de clase
                                naam_row = None
                                for check_row in range(row_idx + 1, min(row_idx + 10, ws.max_row + 1)):
                                    naam_cell = ws.cell(row=check_row, column=1).value
                                    if naam_cell and ('naam' in str(naam_cell).lower()):
                                        naam_row = check_row + 1
                                        break
                                
                                if naam_row:
                                    # Leer registros de estudiantes desde la fila después de "Naam?"
                                    for student_row in range(naam_row, min(naam_row + 50, ws.max_row + 1)):
                                        student_name_cell = ws.cell(row=student_row, column=1).value
                                        if not student_name_cell:
                                            continue
                                        
                                        student_name = str(student_name_cell).strip()
                                        
                                        # Si encontramos otra clase, parar
                                        if re.match(r'^\d+(\/\d+)?\s+[A-Za-z]', student_name):
                                            break
                                        
                                        # Buscar el estudiante en nuestro mapa
                                        student = None
                                        for s in klas_students:
                                            if s['name'].lower() == student_name.lower():
                                                student = s
                                                break
                                        
                                        if student:
                                            # Leer horas 1-7 (columnas B-H = 2-8)
                                            for hour in range(1, 8):
                                                col_idx = hour + 1  # Columna B = 2, C = 3, etc.
                                                cell_value = ws.cell(row=student_row, column=col_idx).value
                                                chillout = parse_chillout_value(cell_value)
                                                if chillout:
                                                    if student['id'] not in daily_record['entries']:
                                                        daily_record['entries'][student['id']] = {}
                                                    daily_record['entries'][student['id']][hour] = chillout
                                    
                                    # También buscar en la sección derecha (columnas K-Q = 11-17)
                                    # Buscar si hay otra sección de clase en la columna K
                                    for check_row in range(row_idx, min(row_idx + 50, ws.max_row + 1)):
                                        klas_cell_right = ws.cell(row=check_row, column=11).value  # Columna K
                                        if klas_cell_right:
                                            klas_str_right = str(klas_cell_right).strip()
                                            if klas_str_right == klas_str:
                                                # Buscar "Naam?" en la columna K
                                                for check_row2 in range(check_row + 1, min(check_row + 10, ws.max_row + 1)):
                                                    naam_cell_right = ws.cell(row=check_row2, column=11).value
                                                    if naam_cell_right and ('naam' in str(naam_cell_right).lower()):
                                                        naam_row_right = check_row2 + 1
                                                        # Leer estudiantes de la sección derecha
                                                        for student_row in range(naam_row_right, min(naam_row_right + 50, ws.max_row + 1)):
                                                            student_name_cell = ws.cell(row=student_row, column=11).value
                                                            if not student_name_cell:
                                                                continue
                                                            
                                                            student_name = str(student_name_cell).strip()
                                                            if re.match(r'^\d+(\/\d+)?\s+[A-Za-z]', student_name):
                                                                break
                                                            
                                                            student = None
                                                            for s in klas_students:
                                                                if s['name'].lower() == student_name.lower():
                                                                    student = s
                                                                    break
                                                            
                                                            if student:
                                                                # Leer horas 1-7 (columnas K-Q = 11-17)
                                                                for hour in range(1, 8):
                                                                    col_idx = hour + 10  # Columna K = 11, L = 12, etc.
                                                                    cell_value = ws.cell(row=student_row, column=col_idx).value
                                                                    chillout = parse_chillout_value(cell_value)
                                                                    if chillout:
                                                                        if student['id'] not in daily_record['entries']:
                                                                            daily_record['entries'][student['id']] = {}
                                                                        # Solo agregar si no existe ya en la izquierda
                                                                        if hour not in daily_record['entries'][student['id']]:
                                                                            daily_record['entries'][student['id']][hour] = chillout
                                                        break
                                                break
                
                records[date_str] = daily_record
                print(f"  Procesada {sheet_name}: {len(daily_record['entries'])} estudiantes con registros")
                
            except Exception as e:
                print(f"Error procesando hoja {sheet_name}: {e}")
                import traceback
                traceback.print_exc()
                continue
        
        wb.close()
        excel_file.close()
        return records
    except Exception as e:
        print(f"Error importando registros diarios: {e}")
        import traceback
        traceback.print_exc()
        return {}

def main():
    # Intentar ambos archivos
    filenames = ["FUNCIONANDO 30-11.xlsx", "Chill-outs New.xlsx"]
    filename = None
    
    for fn in filenames:
        if os.path.exists(fn):
            filename = fn
            break
    
    if not filename:
        print("No se encontró ningún archivo Excel")
        print("Archivos buscados:", filenames)
        return
    
    print(f"Importando datos de {filename}...")
    print("=" * 60)
    
    # Importar estudiantes
    print("\n1. Importando estudiantes...")
    students = import_students(filename)
    print(f"   ✓ Importados {len(students)} estudiantes")
    
    # Crear mapa de estudiantes por nombre para búsqueda rápida
    students_map = {s['name']: s for s in students}
    
    # Mostrar resumen por clase
    klassen = {}
    for student in students:
        klas = student['klas']
        if klas not in klassen:
            klassen[klas] = []
        klassen[klas].append(student['name'])
    
    print(f"\n   Clases encontradas: {len(klassen)}")
    for klas, names in sorted(klassen.items()):
        print(f"   - {klas}: {len(names)} estudiantes")
    
    # Importar registros diarios
    print("\n2. Importando registros diarios...")
    daily_records = import_daily_records(filename, students_map)
    print(f"   ✓ Importadas {len(daily_records)} fechas con registros")
    
    # Contar total de entradas
    total_entries = sum(len(record['entries']) for record in daily_records.values())
    print(f"   ✓ Total de registros de estudiantes: {total_entries}")
    
    # Crear estructura de datos para la app
    app_data = {
        'students': students,
        'dailyRecords': daily_records,
        'weeklyTotals': {}
    }
    
    # Guardar como JSON
    output_file = 'imported_data.json'
    with open(output_file, 'w', encoding='utf-8') as f:
        json.dump(app_data, f, indent=2, ensure_ascii=False)
    
    print("\n" + "=" * 60)
    print(f"\n✓ Datos exportados a {output_file}")
    print(f"\nResumen:")
    print(f"  - Estudiantes: {len(students)}")
    print(f"  - Clases: {len(klassen)}")
    print(f"  - Días con registros: {len(daily_records)}")
    print(f"\nPara importar en la aplicación web:")
    print(f"  1. Ve a http://localhost:3000/import")
    print(f"  2. O copia el contenido de {output_file} y pégalo en la página de importación")

if __name__ == '__main__':
    main()
